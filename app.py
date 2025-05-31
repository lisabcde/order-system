
import os
from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'your_secret_key'

PRODUCTS = [
    {'name': '可樂', 'price': 30},
    {'name': '雪碧', 'price': 30},
    {'name': '礦泉水', 'price': 20},
    {'name': '檸檬茶', 'price': 35},
    {'name': '黑咖啡', 'price': 40}
]

USERS = {
    'de': 'lisa31126',
    'store2': 'password2'
}

if not os.path.exists('orders.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.append(['店家', '商品', '數量', '總金額', '備註'])
    wb.save('orders.xlsx')

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in USERS and USERS[username] == password:
            session['username'] = username
            return redirect(url_for('products'))
        else:
            return render_template('login.html', error='帳號或密碼錯誤')
    return render_template('login.html')

@app.route('/products', methods=['GET', 'POST'])
def products():
    if 'username' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        selected = request.form.getlist('product')
        quantities = request.form.getlist('quantity')
        remark = request.form['remark']

        total = 0
        order_details = []
        for i in range(len(selected)):
            product_name = selected[i]
            quantity = int(quantities[i])
            price = next(p['price'] for p in PRODUCTS if p['name'] == product_name)
            total += price * quantity
            order_details.append(f"{product_name} x {quantity}")

        wb = load_workbook('orders.xlsx')
        ws = wb.active
        ws.append([session['username'], '; '.join(order_details), sum(map(int, quantities)), total, remark])
        wb.save('orders.xlsx')

        send_email(session['username'], order_details, total)

        return render_template('success.html', total=total)

    return render_template('products.html', products=PRODUCTS)

def send_email(user, details, total):
    sender = 'your_gmail@gmail.com'
    receiver = 'gm31126@gmail.com'
    password = 'your_app_password'

    message = MIMEMultipart()
    message['From'] = sender
    message['To'] = receiver
    message['Subject'] = f"{user} 的新訂單"

    body = f"訂單內容：\n" + '\n'.join(details) + f"\n總金額：{total} 元"
    message.attach(MIMEText(body, 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, receiver, message.as_string())
        server.quit()
        print("Email 已發送")
    except Exception as e:
        print(f"Email 發送失敗: {e}")

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
